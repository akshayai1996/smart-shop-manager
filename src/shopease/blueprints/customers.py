from flask import Blueprint, render_template, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime, timedelta
import random

customers_bp = Blueprint('customers', __name__)

CUSTOMERS_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'customers.xlsx')

DUMMY_CUSTOMERS = [
    ("Aarav Sharma",    "9876543210", 5,  4250.00, "1002202609301", "2026-02-10"),
    ("Priya Patel",     "9823456789", 3,  1890.00, "1202202611151", "2026-02-12"),
    ("Rohan Mehta",     "9712345678", 8,  6720.00, "0102202614201", "2026-02-01"),
    ("Sneha Iyer",      "9654321098", 2,   980.00, "1402202610451", "2026-02-14"),
    ("Vikram Singh",    "9543210987", 6,  5100.00, "0802202616301", "2026-02-08"),
    ("Ananya Reddy",    "9432109876", 4,  3200.00, "1102202613001", "2026-02-11"),
    ("Karan Gupta",     "9321098765", 7,  5890.00, "0502202611451", "2026-02-05"),
    ("Meera Nair",      "9210987654", 1,   450.00, "1502202609151", "2026-02-15"),
    ("Arjun Joshi",     "9109876543", 9,  7650.00, "2801202617001", "2026-01-28"),
    ("Divya Kapoor",    "9098765432", 3,  2340.00, "1302202612301", "2026-02-13"),
    ("Rahul Verma",     "8987654321", 5,  4100.00, "0702202615151", "2026-02-07"),
    ("Pooja Desai",     "8876543210", 2,  1560.00, "1602202610001", "2026-02-16"),
    ("Amit Kumar",      "8765432109", 11, 9200.00, "2501202618301", "2026-01-25"),
    ("Nisha Agarwal",   "8654321098", 4,  3450.00, "0902202614451", "2026-02-09"),
    ("Suresh Pillai",   "8543210987", 6,  5600.00, "0302202616001", "2026-02-03"),
    ("Kavya Rao",       "8432109876", 2,  1200.00, "1702202611301", "2026-02-17"),
    ("Deepak Mishra",   "8321098765", 8,  6800.00, "3001202613451", "2026-01-30"),
    ("Sunita Tiwari",   "8210987654", 3,  2700.00, "0602202610151", "2026-02-06"),
    ("Rajesh Pandey",   "8109876543", 5,  4500.00, "0402202617001", "2026-02-04"),
    ("Lakshmi Bhat",    "8098765432", 7,  6100.00, "0202202612301", "2026-02-02"),
]

def ensure_customers_file():
    """Create and seed the customers Excel file if it doesn't exist."""
    if os.path.exists(CUSTOMERS_FILE):
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    # Header styling
    header_fill = PatternFill("solid", fgColor="1A237E")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["Customer Name", "Phone", "Total Orders", "Total Spent (₹)", "Last Invoice No", "Last Visit"]
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 14

    for row_data in DUMMY_CUSTOMERS:
        ws.append(list(row_data))
    
    wb.save(CUSTOMERS_FILE)


def load_customers():
    """Load all customers from Excel into a list of dicts."""
    ensure_customers_file()
    wb = openpyxl.load_workbook(CUSTOMERS_FILE)
    ws = wb.active
    customers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            customers.append({
                'name':       row[0],
                'phone':      str(row[1]),
                'orders':     row[2] or 0,
                'total_spent': row[3] or 0.0,
                'last_invoice': row[4],
                'last_visit': row[5],
            })
    return customers


def upsert_customer(name, phone, amount, invoice_no):
    """Add new customer or update existing one (match by name+phone)."""
    ensure_customers_file()
    wb = openpyxl.load_workbook(CUSTOMERS_FILE)
    ws = wb.active
    today = datetime.now().strftime('%Y-%m-%d')
    
    found = False
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value).strip().lower() == name.strip().lower() and \
           str(row[1].value).strip() == str(phone).strip():
            # Update existing
            row[2].value = (row[2].value or 0) + 1
            row[3].value = round((row[3].value or 0) + amount, 2)
            row[4].value = invoice_no
            row[5].value = today
            found = True
            break
    
    if not found:
        ws.append([name, phone, 1, round(amount, 2), invoice_no, today])
    
    wb.save(CUSTOMERS_FILE)


@customers_bp.route('/customers')
def index():
    customers = load_customers()
    total_customers = len(customers)
    total_revenue = sum(c['total_spent'] for c in customers)
    top_customer = max(customers, key=lambda c: c['total_spent']) if customers else None
    return render_template('customers.html',
                           customers=customers,
                           total_customers=total_customers,
                           total_revenue=total_revenue,
                           top_customer=top_customer)


@customers_bp.route('/customers/download')
def download():
    ensure_customers_file()
    return send_file(CUSTOMERS_FILE, as_attachment=True, download_name='customers.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
